import openpyxl
import xlsxwriter
import numpy as np
import csv
import sys
import os
from datetime import date
def getDate():
    today = date.today()
    return str(today.month) + "_" + str(today.day) + "_" + str(today.year)

date = getDate()
moduleAB = []
moduleA = []
moduleB = []
moduleBprime = []
moduleC = []
moduleD = []

def createModule(sheet):
    matrix = []
    m_row = sheet.max_row
    for i in range(2,m_row+1):
        p1 = sheet.cell(row = i, column = 1).value
        p2 = sheet.cell(row=i, column=2).value
        p3 = sheet.cell(row=i, column=3).value
        p4 = sheet.cell(row=i, column=4).value
        p5 = sheet.cell(row=i, column=5).value
        p6 = sheet.cell(row=i, column=6).value
        p7 = sheet.cell(row=i, column=7).value
        p8 = sheet.cell(row=i, column=8).value
        p9 = sheet.cell(row=i, column=9).value
        p10 = sheet.cell(row=i, column=10).value
        p11 = sheet.cell(row=i, column=11).value
        matrix.append([p1,p2,p3,p4,p5,p6,p7,p8,p9,p10,p11])
    return matrix
def getAbundance(modules):
    matrix = []
    for a in range(len(modules)):
        check = True
        for b in range(len(matrix)):
            if(modules[a][9] == matrix[b][0] and modules[a][10] == matrix[b][1]):
                matrix[b][2] += 1
                check = False
        if(check):
            matrix.append([modules[a][9],modules[a][10],1])
    matrix.sort(key=lambda x: x[2], reverse=True)
    return matrix
def writeSheet(modules, abund, name):
    output = workbook.add_worksheet(name)
    output.write(0, 0, "PDB")
    output.write(0, 1, "Chain res i (1-4)")
    output.write(0, 2, "Chain res j (2-3)")
    output.write(0, 3, "Pos 1")
    output.write(0, 4, "Pos 4")
    output.write(0, 5, "Pos 2")
    output.write(0, 6, "Pos 3")
    output.write(0, 7, "1:4")
    output.write(0, 8, "2:3")
    output.write(0, 9, "1:4 int")
    output.write(0, 10, "2:3 int")
    output.write(0, 12, "Interactions Observed")
    output.write(0, 13, "Interaction Abundance")
    output.write(0, 14, "Interaction Frequency")
    output.set_column(12, 14, 20)

    for x in range(len(modules)):
        output.write(x + 1, 0, modules[x][0])
        output.write(x + 1, 1, modules[x][1])
        output.write(x + 1, 2, modules[x][2])
        output.write_number(x + 1, 3, modules[x][3])
        output.write_number(x + 1, 4, modules[x][4])
        output.write_number(x + 1, 5, modules[x][5])
        output.write_number(x + 1, 6, modules[x][6])
        output.write(x + 1, 7, str(modules[x][7]))
        output.write(x + 1, 8, str(modules[x][8]))
        output.write(x + 1, 9, str(modules[x][9]))
        output.write(x + 1, 10, str(modules[x][10]))
    for y in range(len(abund)):
        freq = abund[y][2] / len(modules)
        output.write(y + 1, 12, str(abund[y][0] + "/" + abund[y][1]))
        output.write_number(y + 1, 13, abund[y][2])
        output.write_number(y + 1, 14, freq)
def makeTable(modules, abund, name):
    output = workbook.add_worksheet(name)
    x = 0
    for q in range(len(abund)):
        output.write(x,0, abund[q][0] + "/" + abund[q][1])
        output.write(x, 20, abund[q][0] + "/" + abund[q][1] + "Percentages")
        matrix = []
        row = ["A-A", "A-C", "A-G", "A-U", "C-A", "C-C", "C-G", "C-U","G-A", "G-C", "G-G", "G-U","U-A", "U-C", "U-G", "U-U"]
        for line in range(len(row)):
            matrix.append([0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
        count = 0
        for i in range(len(modules)):
            if(modules[i][9] == abund[q][0] and modules[i][10] == abund[q][1]):
                for a in range(len(row)):
                    for b in range(len(row)):
                        if(modules[i][7] == row[b] and modules[i][8] == row[a]):
                            matrix[a][b] += 1
                            count += 1
        matrix = sumTotal(matrix)

        row.append("Total")
        writeTables(matrix, x+1, row, output,count)
        x += 21
def sumTotal(matrix):
    for a in range(len(matrix)):
        sum = 0
        for b in range(len(matrix[a])):
            sum += matrix[a][b]
        matrix[a].append(sum)
    list = []
    for y in range(len(matrix[0])):
        sum = 0
        for x in range(len(matrix)):
            sum += matrix[x][y]
        list.append(sum)
    matrix.append(list)
    return matrix

def writeTables(matrix, x, row, output,count):
    output.add_table(x, 0, x + 17, 17, {'header_row': True, 'autofilter': False, 'first_column': True })
    output.write(x, 0, "n = " + str(count))
    output.write_row(x, 1, row)
    output.write_column(x + 1, 0, row)

    yellow = workbook.add_format({'bg_color': '#FFFF00'})
    for a in range(len(matrix)):
        for b in range(len(matrix[0])):
            if(matrix[a][b]>0):
                output.write(x+1+a,b+1,matrix[a][b],yellow)
            else:
                output.write(x+1+a,b+1,matrix[a][b])
    cell_format = workbook.add_format()
    cell_format.set_num_format(10)
    cell_format1 = workbook.add_format({'bg_color': 'yellow', 'num_format': '0.0%'})

    for w in range(len(matrix)):
        for z in range(len(matrix[0])):
            if(count > 0):
                matrix[w][z] = matrix[w][z] / count

    output.add_table(x, 20, x + 17, 37, {'header_row': True, 'autofilter': False, 'first_column': True})
    output.write(x, 20, "n = " + str(count))
    output.write_row(x, 21, row)
    output.write_column(x + 1, 20, row)

    for a1 in range(len(matrix)):
        for b1 in range(len(matrix[0])):
            if (matrix[a1][b1] > 0):
                output.write(x + 1 + a1, b1 + 21, matrix[a1][b1], cell_format1)
            else:
                output.write(x + 1 + a1, b1 + 21, matrix[a1][b1],cell_format)

for file in range(1,len(sys.argv)):
    wb = openpyxl.load_workbook(sys.argv[file])
    moduleAB += (createModule(wb['Output A+B+Bprime']))
    moduleA += (createModule(wb["OutputA"]))
    moduleB += (createModule(wb["OutputB"]))
    moduleBprime += (createModule(wb['Output B-prime']))
    moduleC += (createModule(wb["OutputC"]))
    moduleD += (createModule(wb["OutputD"]))

abAbun = getAbundance(moduleAB)
aAbun = getAbundance(moduleA)
bAbun = getAbundance(moduleB)
bpAbun = getAbundance(moduleBprime)
cAbun = getAbundance(moduleC)
dAbun = getAbundance(moduleD)



file_name = "Modules_Combined_" + date + ".xlsx"
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file_path = os.path.join(desktop_path, file_name)
workbook = xlsxwriter.Workbook(file_path)

writeSheet(moduleAB, abAbun, "Output A+B+Bprime")
makeTable(moduleAB, abAbun, "Tables Output A+B+Bprime")
writeSheet(moduleA, aAbun, "OutputA")
makeTable(moduleA, aAbun, "Tables Output A")
writeSheet(moduleB, bAbun, "OutputB")
makeTable(moduleB, bAbun, "Tables Output B")
writeSheet(moduleBprime, bpAbun, "Output B-prime")
makeTable(moduleBprime, bpAbun, "Tables Output B-prime")
writeSheet(moduleC, cAbun, "OutputC")
makeTable(moduleC, cAbun, "Tables Output C")
writeSheet(moduleD, dAbun, "OutputD")
makeTable(moduleD, dAbun, "Tables Output D")

workbook.close()

print(f"Excel file '{file_name}' created on your desktop.")
print("Program Complete")