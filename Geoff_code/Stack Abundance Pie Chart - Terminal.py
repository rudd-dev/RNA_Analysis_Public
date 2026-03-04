import openpyxl
import xlsxwriter
import numpy as np
import csv
import sys
import os
from datetime import date

class Stack():
    base1 = ""
    base2 = ""
    int1 = ""
    int2 = ""

    def __init__(self, b1,b2,i1,i2):
        self.base1 = b1
        self.base2 = b2
        self.int1 = i1
        self.int2 = i2

    def getBase1(self):
        return self.base1
    def getBase2(self):
        return self.base2
    def getInt1(self):
        return self.int1
    def getInt2(self):
        return self.int2
    def getInt(self):
        return str(self.int1) + "/" + str(self.int2)
    def getBase(self):
        return str(self.base1) + "/" + str(self.base2)
    def __str__(self):
        return str(self.base1) + "/" + str(self.base2) + " - " + str(self.int1) + "/" + str(self.int2)

def getDate():
    today = date.today()
    return str(today.month) + "_" + str(today.day) + "_" + str(today.year)
def createModule(sheet):
    list = []
    m_row = sheet.max_row
    for i in range(2,m_row+1):
        p8 = sheet.cell(row=i, column=8).value
        p9 = sheet.cell(row=i, column=9).value
        p10 = sheet.cell(row=i, column=10).value
        p11 = sheet.cell(row=i, column=11).value
        list.append(Stack(p8,p9,p10,p11))
    return list

def has_target_codes(s):
    targets = ["cWW", "tSR", "tRS", "tSS", "cSR", "cRS"]
    return any(code in s for code in targets)

def createWHcol(sheet):
    list = ['cWW/cWW','x/cWW','cWW/x']
    x = 2
    while(len(list)<21):
        if(x>450):
            break
        int = str(sheet.cell(row=x, column=13).value)
        if(has_target_codes(int) == False):
            list.append(str(int))
        x+=1
    return list

def createPairs():
    pairs = ["C-G", "G-C", "A-U", "U-A", "U-G", "G-U", "A-G", "G-A", "A-A", "A-C", "C-A", "C-C", "C-U", "G-G", "U-C",
             "U-U"]
    final = []
    for a in range(len(pairs)):
        for b in range(len(pairs)):
            final.append(pairs[a] + "/" + pairs[b])
    return final

def getReverse(str):
    return str[0:1] + str[2:3] + str[1:2]


def columnExtract(matrix,x):
    list = []
    for i in range(len(matrix)):
        list.append(matrix[i][x])
    return list

def leftTotal(mod,line):
    total = 0
    for b in range(len(mod)):
        if(line == mod[b].getBase()):
            total += 1
    return total

def findcWW(mod,line):
    total = 0
    for b in range(len(mod)):
        if(line == mod[b].getBase()):
            if (mod[b].getInt() == 'cWW/cWW'):
                total += 1
    return total

def findSS(mod,line):
    total = 0
    for b in range(len(mod)):
        if (line == mod[b].getBase()):
            if(mod[b].getInt1() == 'tSS' or mod[b].getInt2() == 'tSS' or mod[b].getInt1() == 'cSS' or mod[b].getInt2() == 'cSS' or mod[b].getInt1() == 'tSR' or mod[b].getInt2() == 'tSR' or mod[b].getInt1() == 'cSR' or mod[b].getInt2() == 'cSR' or mod[b].getInt1() == 'tRS' or mod[b].getInt2() == 'tRS' or mod[b].getInt1() == 'cRS' or mod[b].getInt2() == 'cRS'):
                total += 1
    return total

def findXcWW(mod,line):
    total = 0
    for c in range(len(mod)):
        if (line == mod[c].getBase()):
            if (mod[c].getInt1() != 'cWW' and mod[c].getInt2() == 'cWW'):
                if (noSug(mod[c].getInt1())):
                    total += 1
    return total

def findcWWX(mod,line):
    total = 0
    for c in range(len(mod)):
        if (line == mod[c].getBase()):
            if (mod[c].getInt2() != 'cWW' and mod[c].getInt1() == 'cWW'):
                if (noSug(mod[c].getInt2())):
                    total += 1
    return total

def getA(mod, pair): #finds the occurence of cWW/cWW given the nucleotide pair
    count = 0
    for i in mod:
        if(i.getBase() == pair):
            if(i.getInt() == "cWW/cWW"):
                count+=1
    return count

def getnA(mod, pair): #finds the occurence of anything that is not cWW/cWW given the nucleotide pair
    count = 0
    for i in mod:
        if(i.getBase() == pair):
            if(i.getInt() != "cWW/cWW"):
                count+=1
    return count
def writeSheetA(mod):
    output = workbook.add_worksheet("Output 2H")
    pairs = createPairs()
    header = ["cWW","Other","Total"]
    output.write_row(0, 1, header)
    output.write_column(1, 0, pairs)
    list = []
    for a in range(len(pairs)):
        output.write(a+1, 1, int(getA(mod, pairs[a])))
        output.write(a + 1, 2, int(getnA(mod, pairs[a])))
        list.append(int(getA(mod, pairs[a]))+int(getnA(mod, pairs[a])))
    output.write_column(1, 3, list)
    return list

def writeSheetAA(modA,modB,modBp,mod):
    output = workbook.add_worksheet("Output 2H Split")
    pairs = createPairs()
    header = ["A-cWW","B-cWW","Bp-cWW","A-Other","B-Other","Bp-Other","Total"]
    output.write_row(0, 1, header)
    output.write_column(1, 0, pairs)
    list = []
    for a in range(len(pairs)):
        output.write(a + 1, 1, int(getA(modA, pairs[a])))
        output.write(a + 1, 2, int(getA(modB, pairs[a])))
        output.write(a + 1, 3, int(getA(modBp, pairs[a])))
        output.write(a + 1, 4, int(getnA(modA, pairs[a])))
        output.write(a + 1, 5, int(getnA(modB, pairs[a])))
        output.write(a + 1, 6, int(getnA(modBp, pairs[a])))
        list.append(int(getA(mod, pairs[a]))+int(getnA(mod, pairs[a])))
    output.write_column(1, 7, list)
    return list

def writeSheetAA2H(modA,modB,modBp,mod):
    output = workbook.add_worksheet("Output Only 2H Split")
    pairs = createPairs()
    header = ["A-cWW","B-cWW","Bp-cWW","Total"]
    output.write_row(0, 1, header)
    output.write_column(1, 0, pairs)
    list = []
    for a in range(len(pairs)):
        output.write(a + 1, 1, int(getA(modA, pairs[a])))
        output.write(a + 1, 2, int(getA(modB, pairs[a])))
        output.write(a + 1, 3, int(getA(modBp, pairs[a])))
        list.append(int(getA(mod, pairs[a])))
    output.write_column(1, 4, list)
    return list

def writeSheetAAn2H(modA,modB,modBp,mod):
    output = workbook.add_worksheet("Output non 2H Split")
    pairs = createPairs()
    header = ["A-Other","B-Other","Bp-Other","Total"]
    output.write_row(0, 1, header)
    output.write_column(1, 0, pairs)
    list = []
    for a in range(len(pairs)):
        output.write(a + 1, 1, int(getnA(modA, pairs[a])))
        output.write(a + 1, 2, int(getnA(modB, pairs[a])))
        output.write(a + 1, 3, int(getnA(modBp, pairs[a])))
        list.append(int(getnA(mod, pairs[a])))
    output.write_column(1, 4, list)
    return list

def noSug(str):
    for i in range(len(str)):
        str = str.lower()
        return "sr" not in str and "rs" not in str and "ss" not in str

def getB(line,header,mod):
    matrix = []
    for a in range(len(line)):
        list = []
        total = 0
        non = 0
        for b in range(len(header)-2):
            count = 0
            for c in range(len(mod)):
                if(mod[c].getBase() == line[a]):
                    if(mod[c].getInt1() == "cWW" and mod[c].getInt2() == header[b]):
                        count += 1
                        non +=1
                    if (noSug(mod[c].getInt1()) and noSug(mod[c].getInt2())):
                        total += 1
            list.append(count)
        total /= 16
        list.append(total - non)
        list.append(total)
        matrix.append(list)
    return matrix
def writeSheetB(mod):
    output = workbook.add_worksheet("Output cWW-x")
    pairs = createPairs()
    col = ['cWW','tWH','cSW',"cSH","tSW",'tWW','cWH','cHW','tSH','tHW','tHS','tHH','cHS','cWS','cHH','tWS','Other','Total']
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getB(pairs,col,mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i+1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list
def getC(line,header,mod):
    matrix = []
    for a in range(len(line)):
        total = 0
        non = 0
        list = []
        for b in range(len(header)-2):
            count = 0
            for c in range(len(mod)):
                if(mod[c].getBase() == line[a]):
                    if(mod[c].getInt2() == "cWW" and mod[c].getInt1() == header[b]):
                        count += 1
                        non += 1
                    if(noSug(mod[c].getInt1()) and noSug(mod[c].getInt2())):
                        total+=1
            list.append(count)
        total /= 16
        list.append(total-non)
        list.append(total)
        matrix.append(list)
    return matrix

def writeSheetC(mod):
    output = workbook.add_worksheet("Output x-cWW")
    pairs = createPairs()
    col = ['cWW','cSW','cHS','cHW','cWS','tWW','tSW','tSH','tHW','cWH','tWS','tHS','tHH','cSH','tWH','cHH','Other','Total']
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getC(pairs,col,mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i+1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list
def getE(line,header,mod): #other is anything with the SR, RS, or SS
    matrix = []
    for a in range(len(line)):
        non = 0
        total = 0
        list = []
        for b in range(len(header)-2):
            count = 0
            for c in range(len(mod)):
                if(mod[c].getBase() == line[a]):
                    if(mod[c].getInt1() == 'tSS' or mod[c].getInt2() == 'tSS' or mod[c].getInt1() == 'cSR' or mod[c].getInt2() == 'cSR'or mod[c].getInt1() == 'cRS' or mod[c].getInt2() == 'cRS' or mod[c].getInt1() == 'tSR' or mod[c].getInt2() == 'tSR' or mod[c].getInt1() == 'tRS' or mod[c].getInt2() == 'tRS'):
                        total += 1
                    if(mod[c].getInt() == header[b]):
                        count += 1
                        non += 1
            list.append(count)
        total /= 35
        list.append(total-non)
        list.append(total)
        matrix.append(list)
    return matrix
def writeSheetE(mod):
    output = workbook.add_worksheet("Output SS")
    pairs = createPairs()
    col = ['cWW/tSS' ,'cWW/tRS' ,'cWW/tSR' ,'cWW/cRS' ,'cWW/cSR' ,'tSS/cWW' ,'tRS/cWW' ,'tSR/cWW' ,'cRS/cWW' ,'cSR/cWW' ,'tSS/tSS' ,'tSS/tRS' ,'tSS/tSR' ,'tSS/cRS' ,'tSS/cSR' ,'tRS/tSS' ,'tRS/tRS' ,'tRS/tSR' ,'tRS/cRS' ,'tRS/cSR' ,'tSR/tSS' ,'tSR/tRS' ,'tSR/tSR' ,'tSR/cRS' ,'tSR/cSR' ,'cRS/tSS' ,'cRS/tRS' ,'cRS/tSR' ,'cRS/cRS' ,'cRS/cSR' ,'cSR/tSS' ,'cSR/tRS' ,'cSR/tSR' ,'cSR/cRS' ,'cSR/cSR','Other','Total']
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getE(pairs,col,mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i+1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list
def getAE(line,header,mod):
    matrix = []
    for a in range(len(line)):
        cWW = 0
        total = 0
        SS = 0
        list = []
        for b in range(len(header)-2):
            count = 0
            for c in range(len(mod)):
                if(mod[c].getBase() == line[a]):
                    total += 1
                    if(mod[c].getInt() == header[b]):
                        count += 1
                    if (mod[c].getInt() == header[b] and header[b] == 'cWW/cWW'):
                        cWW += 1
                    if (mod[c].getInt() == header[b] and header[b] != 'cWW/cWW'):
                        SS += 1
            list.append(count)
        total /= 36
        list.append(total - (cWW + SS))
        list.append(total)
        matrix.append(list)
    return matrix

def writeSheetAE(mod):
    output = workbook.add_worksheet("Output 2H + SS")
    pairs = createPairs()
    col = ['cWW/cWW',"cWW/tSS",'cWW/tRS','cWW/tSR','cWW/cRS','cWW/cSR','tSS/cWW','tRS/cWW','tSR/cWW','cRS/cWW','cSR/cWW','tSS/tSS','tSS/tRS','tSS/tSR','tSS/cRS','tSS/cSR','tRS/tSS','tRS/tRS','tRS/tSR','tRS/cRS','tRS/cSR','tSR/tSS','tSR/tRS','tSR/tSR','tSR/cRS','tSR/cSR','cRS/tSS','cRS/tRS','cRS/tSR','cRS/cRS','cRS/cSR','cSR/tSS','cSR/tRS','cSR/tSR','cSR/cRS','cSR/cSR','Other(W/H)','Total']
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getAE(pairs, col, mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i + 1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list

def getNC(line,header,mod):
    matrix = []
    for a in range(len(line)):
        list = []
        WH = leftTotal(mod, line[a])
        cWW = findcWW(mod,line[a])
        xc = findXcWW(mod,line[a])
        cx = findcWWX(mod,line[a])
        list.append(cWW)
        list.append(xc)
        list.append(cx)
        sum = 0
        for b in range(3,len(header) - 2):
            count = 0
            for c in range(len(mod)):
                if (mod[c].getBase() == line[a]):
                    if (mod[c].getInt() == header[b]):
                        count += 1
                        sum+=1
            list.append(count)
        list.append(WH - (sum + cWW+xc+cx))
        list.append(WH)
        matrix.append(list)
    return matrix

def writeSheetNC(mod,col):
    output = workbook.add_worksheet("Output 2H + NC")
    pairs = createPairs()
    col.append('Other(W/H)')
    col.append("Total")
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getNC(pairs, col, mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i + 1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list

def getDub(line,header,mod):
    matrix = []
    for a in range(len(line)):
        list = []
        sum = 0
        total = 0
        for b in range(len(header) - 2):
            count = 0
            for c in range(len(mod)):
                if (mod[c].getBase() == line[a]):
                    total += 1
                    if (mod[c].getInt() == header[b][0:7]):
                        count += 1
                        sum+=1
            list.append(count)
        list.append((total/13)-sum)
        list.append(total/13)
        matrix.append(list)
    return matrix

def getColPre(a,b):
    wb = openpyxl.load_workbook(sys.argv[1])
    list = []
    other = 0
    total = 0
    for w in range(2,b):
        other += wb['OutputD'].cell(row=w, column=15).value

    for i in range(a, b):
        per = round(wb['OutputD'].cell(row=i, column=15).value * 100, 2)
        value = str(wb['OutputD'].cell(row=i, column=13).value) + '-' + str(per) + "%"
        list.append(value)
    list.append('Other-' + str(round(((1 - other) * 100), 2)) + "%")
    list.append('Total')
    return list

def writeSheetDub(mod):
    output = workbook.add_worksheet("Output Doublets")
    pairs = createPairs()
    col = getColPre(2,15)
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getDub(pairs, col, mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i + 1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list

def getDubNC(line,header,mod):
    matrix = []
    for a in range(len(line)):
        list = []
        sum = 0
        total = 0
        for b in range(len(header) - 2):
            count = 0
            for c in range(len(mod)):
                if (mod[c].getBase() == line[a] and mod[c].getInt() != 'cWW/cWW'):
                    total += 1
                    if (mod[c].getInt() == header[b][0:7]):
                        count += 1
                        sum+=1
            list.append(count)
        list.append((total/17)-sum)
        list.append(total/17)
        matrix.append(list)
    return matrix

def writeSheetDubNC(mod):
    output = workbook.add_worksheet("Output Doublets(no cWC)")
    pairs = createPairs()
    col = getColPre(3,20)
    output.write_row(0, 1, col)
    output.write_column(1, 0, pairs)
    matrix = getDubNC(pairs, col, mod)
    list = []
    for i in range(len(pairs)):
        output.write_row(i + 1, 1, matrix[i])
        list.append(matrix[i][-1])
    return list

def writeGraphA(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output 2H Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output 2H", 0, 1, 0, 2],
                'values': ["Output 2H", a+1, 1, a+1, 2],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphAA(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output 2H Split Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'categories': ["Output 2H Split", 0, 1, 0, 6],
            'values': ["Output 2H Split", a + 1, 1, a + 1, 6],
            'data_labels': {'percentage': True},
        })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphAA2H(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output Only 2H Split Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'categories': ["Output Only 2H Split", 0, 1, 0, 3],
            'values': ["Output Only 2H Split", a + 1, 1, a + 1, 3],
            'data_labels': {'percentage': True},
        })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0
def writeGraphAAn2H(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output non 2H Split Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'categories': ["Output non 2H Split", 0, 1, 0, 3],
            'values': ["Output non 2H Split", a + 1, 1, a + 1, 3],
            'data_labels': {'percentage': True},
        })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphB(list):
    bp = ['n=10',"A-A", "A-C", "A-G", "A-U", "C-A", "C-C", "C-G", "C-U", "G-A", "G-C", "G-G", "G-U", "U-A", "U-C", "U-G",
             "U-U"]
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output cWW-x Charts")
    '''format = workbook.add_format({'font_size': 30, 'align': 'center'})
    worksheets.write_column(0,0,bp,format)
    worksheets.write_row(0, 0, bp,format)'''
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output cWW-x", 0, 1, 0, 17],
                'values': ["Output cWW-x", a+1, 1, a+1, 17],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphC(list):
    bp = ['n=10',"A-A", "A-C", "A-G", "A-U", "C-A", "C-C", "C-G", "C-U", "G-A", "G-C", "G-G", "G-U", "U-A", "U-C", "U-G",
             "U-U"]
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output x-cWW Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output x-cWW", 0, 1, 0, 17],
                'values': ["Output x-cWW", a+1, 1, a+1, 17],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0
def writeGraphE(list):
    bp = ['n=10',"A-A", "A-C", "A-G", "A-U", "C-A", "C-C", "C-G", "C-U", "G-A", "G-C", "G-G", "G-U", "U-A", "U-C", "U-G",
             "U-U"]
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output SS Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output SS", 0, 1, 0, 36],
                'values': ["Output SS", a+1, 1, a+1, 36],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y += 15
        count += 1
        if (count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphAE(list):
    bp = ['n=10',"A-A", "A-C", "A-G", "A-U", "C-A", "C-C", "C-G", "C-U", "G-A", "G-C", "G-G", "G-U", "U-A", "U-C", "U-G",
             "U-U"]
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output 2H + SS Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output 2H + SS", 0, 1, 0, 37],
                'values': ["Output 2H + SS", a+1, 1, a+1, 37],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y+=15
        count += 1
        if(count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphNC(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output 2H + NC Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output 2H + NC", 0, 1, 0, 23],
                'values': ["Output 2H + NC", a+1, 1, a+1, 23],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y+=15
        count += 1
        if(count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphDub(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output Doublets Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output Doublets", 0, 1, 0, 14],
                'values': ["Output Doublets", a+1, 1, a+1, 14],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y+=15
        count += 1
        if(count == 16):
            count = 0
            x += 8
            y = 0

def writeGraphDubNC(list):
    pairs = createPairs()
    worksheets = workbook.add_worksheet("Output Doublets(no cWC) Charts")
    x = 0
    y = 0
    count = 0
    for a in range(len(pairs)):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
                'categories': ["Output Doublets(no cWC)", 0, 1, 0, 18],
                'values': ["Output Doublets(no cWC)", a+1, 1, a+1, 18],
                'data_labels': {'percentage': True},
            })
        chart.set_title({'name': pairs[a] + " - n=" + str(int(list[a]))})
        chart.set_legend({'position': 'bottom'})
        worksheets.insert_chart(y, x, chart)
        y+=15
        count += 1
        if(count == 16):
            count = 0
            x += 8
            y = 0
def sumTotal(matrix):
    for a in range(len(matrix)):
        sum = 0
        for b in range(len(matrix[a])):
            sum += matrix[a][b]
        matrix[a].append(sum)
    list = []
    for y in range(len(matrix[0])):
        sum = 0
        for x in range(len(matrix)):
            sum += matrix[x][y]
        list.append(sum)
    matrix.append(list)
    return matrix
def makeTable(modules,name):
    pairs = createPairs() #'C-G/G-C
    output = workbook.add_worksheet("Tables " + str(name))
    x = 0
    y = 0
    for q in range(len(pairs)):
        row = ['cWW', 'cWH', 'cWS', 'cHH', 'cHW', 'cHS', 'cRS','cSR', 'cSW', 'cSH', 'tWW', 'tWH', 'tWS', 'tHH', 'tHW', 'tHS',
               'tSS','tSR','tRS', 'tSW', 'tSH']
        output.write(x, y, pairs[q])
        output.write(x, y+25, pairs[q] + " Percentages")
        matrix = []
        for line in range(len(row)):
            matrix.append([0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
        count = 0
        for i in range(len(modules)):
            if (modules[i].getBase() == pairs[q]):
                for a in range(len(row)):
                    for b in range(len(row)):
                        if (modules[i].getInt1() == row[b] and modules[i].getInt2() == row[a]):
                            matrix[a][b] += 1
                            count += 1
        matrix = sumTotal(matrix)
        row.append("Total")
        writeTables(matrix, x + 1, y, output, count)

        x+= 26
        if(x > 400):
            x = 0
            y += 51

def writeTables(matrix, x,y, output,count):
    output.add_table(x, y, x + 22, y+22, {'header_row': True, 'autofilter': False, 'first_column': True })
    row = ['cWW', 'cWH', 'cWS', 'cHH', 'cHW', 'cHS', 'cRS','cSR', 'cSW', 'cSH', 'tWW', 'tWH', 'tWS', 'tHH', 'tHW', 'tHS',
               'tSS','tSR','tRS', 'tSW', 'tSH','Total']
    output.write(x, y, "n = " + str(count))
    output.write_row(x, y+1, row)
    output.write_column(x + 1, y, row)

    yellow = workbook.add_format({'bg_color': '#FFFF00'})
    for a in range(len(matrix)):
        for b in range(len(matrix[0])):
            if(matrix[a][b]>0):
                output.write(x+1+a,y+b+1,matrix[a][b],yellow)
            else:
                output.write(x+1+a,y+b+1,matrix[a][b])
    cell_format = workbook.add_format()
    cell_format.set_num_format(10)
    cell_format1 = workbook.add_format({'bg_color': 'yellow', 'num_format': '0.0%'})

    for w in range(len(matrix)):
        for z in range(len(matrix[0])):
            if(count > 0):
                matrix[w][z] = matrix[w][z] / count

    output.add_table(x, y+25, x + 22, y+47, {'header_row': True, 'autofilter': False, 'first_column': True})
    output.write(x, y+25, "n = " + str(count))
    output.write_row(x, y+26, row)
    output.write_column(x + 1, y+25, row)

    for a1 in range(len(matrix)):
        for b1 in range(len(matrix[0])):
            if (matrix[a1][b1] > 0):
                output.write(x + 1 + a1, y+b1 + 26, matrix[a1][b1], cell_format1)
            else:
                output.write(x + 1 + a1, y+b1 + 26, matrix[a1][b1],cell_format)

wb = openpyxl.load_workbook(sys.argv[1])
modules = (createModule(wb['Output A+B+Bprime']))
modulesD = createModule(wb['OutputD'])
modA = createModule(wb['OutputA'])
modB = createModule(wb['OutputB'])
modBp = createModule(wb['Output B-prime'])

date = getDate()
file_name = "Stack_Abundance_" + date + ".xlsx"
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file_path = os.path.join(desktop_path, file_name)
workbook = xlsxwriter.Workbook(file_path)


listA = writeSheetA(modules)
writeGraphA(listA)
listAA = writeSheetAA(modA,modB,modBp,modules)
writeGraphAA(listAA)
listAA2H = writeSheetAA2H(modA,modB,modBp,modules)
writeGraphAA2H(listAA2H)
listAAn2H = writeSheetAAn2H(modA,modB,modBp,modules)
writeGraphAAn2H(listAAn2H)
listB = writeSheetB(modules)
writeGraphB(listB)
listC = writeSheetC(modules)
writeGraphC(listC)
listE = writeSheetE(modules)
writeGraphE(listE)
listAE = writeSheetAE(modules)
writeGraphAE(listAE)
listNC = writeSheetNC(modules,createWHcol(wb['Output A+B+Bprime']))
writeGraphNC(listNC)
makeTable(modules, "A+B+Bprime")
listDub = writeSheetDub(modulesD)
writeGraphDub(listDub)
listDubNC = writeSheetDubNC(modulesD)
writeGraphDubNC(listDubNC)
makeTable(modulesD, "Doublets")

workbook.close()

print(f"Excel file '{file_name}' created on your desktop.")
print("Program Complete")



